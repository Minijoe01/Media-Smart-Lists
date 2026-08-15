"""Migration ZIP Trakt → MDBList (assistant web).

Lit un NormalizedDataset issu d'un ZIP Trakt (via trakt_zip_provider),
construit un plan de migration, des payloads MDBList avec les vraies dates
(watched_at), identifie les contenus sans correspondance, et génère un
rapport Excel de ce qui a été chargé / échoué.

Mode simulation (dry-run) : aucun POST n'est envoyé, le rapport est quand
même généré.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ── Helpers ──────────────────────────────────────────────────────────────────


def _has_mapping(ids: dict[str, Any]) -> bool:
    """Un contenu est migrable s'il a un id TMDb ou IMDb utilisable."""
    if not isinstance(ids, dict):
        return False
    return bool(ids.get("tmdb") or ids.get("imdb"))


def _ids_block(ids: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    if ids.get("tmdb"):
        out["tmdb"] = int(ids["tmdb"])
    if ids.get("imdb"):
        out["imdb"] = str(ids["imdb"])
    return out


def _safe_date(value: Any) -> str | None:
    if not value:
        return None
    try:
        parsed = datetime.fromisoformat(str(value).replace("Z", "+00:00"))
        if parsed.tzinfo is None:
            parsed = parsed.replace(tzinfo=timezone.utc)
        return parsed.isoformat()
    except (TypeError, ValueError):
        return None


# ── Plan de migration ────────────────────────────────────────────────────────


def build_migration_plan(dataset: dict[str, Any]) -> dict[str, Any]:
    """Analyse le dataset ZIP et produit le plan : compteurs, contenus sans
    correspondance, et données prêtes pour l'aperçu."""
    sections = dataset.get("sections") or {}
    watched = sections.get("watched") or {}
    ratings = sections.get("ratings") or {}
    watchlist = sections.get("watchlist") or {}
    user_lists = sections.get("user_lists") or []

    # ── Films vus (dernière date) ──
    movies_seen: dict[str, dict[str, Any]] = {}
    for row in watched.get("movies") or []:
        if not isinstance(row, dict):
            continue
        movie = row.get("movie") if isinstance(row.get("movie"), dict) else row
        ids = movie.get("ids") if isinstance(movie.get("ids"), dict) else {}
        key = str(ids.get("tmdb") or ids.get("imdb") or movie.get("title") or "?")
        if not _has_mapping(ids):
            continue
        entry = movies_seen.setdefault(
            key,
            {"title": movie.get("title"), "year": movie.get("year"), "ids": ids,
             "watched_at": _safe_date(row.get("last_watched_at") or row.get("watched_at"))},
        )
        # plays : nombre de visionnages (rewatches)
        entry["plays"] = int(entry.get("plays") or 0) + int(row.get("plays") or 1)
        date_now = _safe_date(row.get("last_watched_at") or row.get("watched_at"))
        if date_now and (not entry["watched_at"] or date_now > entry["watched_at"]):
            entry["watched_at"] = date_now

    # ── Épisodes vus : regrouper par série, saison, épisode avec dates ──
    # (MDBList ne garde qu'une date par épisode : on prend la dernière.)
    episodes_by_show: dict[str, dict[str, Any]] = {}
    for row in watched.get("episodes") or []:
        if not isinstance(row, dict):
            continue
        episode = row.get("episode") if isinstance(row.get("episode"), dict) else row
        show = row.get("show") if isinstance(row.get("show"), dict) else (episode.get("show") if isinstance(episode.get("show"), dict) else {})
        show_ids = show.get("ids") if isinstance(show.get("ids"), dict) else {}
        if not _has_mapping(show_ids):
            continue
        show_key = str(show_ids.get("tmdb") or show_ids.get("imdb") or show.get("title") or "?")
        entry = episodes_by_show.setdefault(
            show_key,
            {"title": show.get("title"), "year": show.get("year"), "ids": show_ids,
             "seasons": {}},
        )
        try:
            season = int(episode.get("season"))
            number = int(episode.get("number"))
        except (TypeError, ValueError):
            continue
        date_now = _safe_date(row.get("last_watched_at") or row.get("watched_at"))
        existing = entry["seasons"].get(season, {}).get(number)
        if not existing or (date_now and date_now > existing):
            entry["seasons"].setdefault(season, {})[number] = date_now or ""

    # ── Contenus SANS correspondance (pas d'id TMDb/IMDb) ──
    no_match: list[dict[str, Any]] = []
    for row in watched.get("movies") or []:
        if not isinstance(row, dict):
            continue
        movie = row.get("movie") if isinstance(row.get("movie"), dict) else row
        ids = movie.get("ids") if isinstance(movie.get("ids"), dict) else {}
        if not _has_mapping(ids):
            no_match.append({"type": "Film", "title": movie.get("title"), "year": movie.get("year"), "reason": "aucun id TMDb/IMDb"})
    for row in watched.get("episodes") or []:
        if not isinstance(row, dict):
            continue
        show = row.get("show") if isinstance(row.get("show"), dict) else {}
        ids = show.get("ids") if isinstance(show.get("ids"), dict) else {}
        if not _has_mapping(ids):
            no_match.append({"type": "Série", "title": show.get("title"), "year": show.get("year"), "reason": "aucun id TMDb/IMDb"})
    # dédupliquer le sans-correspondance
    seen_no = set()
    no_match_unique = []
    for item in no_match:
        marker = (item["type"], str(item["title"]), str(item["year"]))
        if marker not in seen_no:
            seen_no.add(marker)
            no_match_unique.append(item)

    # ── Notes ──
    ratings_out = {
        "movies": len(ratings.get("movies") or []),
        "shows": len(ratings.get("shows") or []),
        "episodes": len(ratings.get("episodes") or []),
    }

    # ── Watchlist ──
    watchlist_out = {
        "movies": len(watchlist.get("movies") or []),
        "shows": len(watchlist.get("shows") or []),
    }

    # ── Listes ──
    lists_out = [
        {"name": item.get("name"), "movies": len(item.get("movies") or []), "shows": len(item.get("shows") or [])}
        for item in user_lists
        if isinstance(item, dict)
    ]

    return {
        "films_vus": len(movies_seen),
        "episodes_vus": sum(len(s) for e in episodes_by_show.values() for s in e["seasons"].values()),
        "series_vues": len(episodes_by_show),
        "rewatches": sum(1 for m in movies_seen.values() if (m.get("plays") or 1) > 1),
        "notes": ratings_out,
        "watchlist": watchlist_out,
        "listes": lists_out,
        "lists_detail": build_lists_plans(dataset),
        "sans_correspondance": no_match_unique,
        "movies_seen": movies_seen,
        "episodes_by_show": episodes_by_show,
    }


# ── Payloads MDBList ─────────────────────────────────────────────────────────


def build_watched_payloads(plan: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    """Construit les payloads POST /sync/watched avec les vraies dates."""
    movies = []
    for entry in plan["movies_seen"].values():
        item = _ids_block(entry["ids"])
        if entry.get("watched_at"):
            item["watched_at"] = entry["watched_at"]
        movies.append(item)

    shows = []
    for entry in plan["episodes_by_show"].values():
        item = _ids_block(entry["ids"])
        seasons = []
        for season_num in sorted(entry["seasons"]):
            episodes = []
            for ep_num, date_str in sorted(entry["seasons"][season_num].items()):
                ep_item: dict[str, Any] = {"number": ep_num}
                if date_str:
                    ep_item["watched_at"] = date_str
                episodes.append(ep_item)
            seasons.append({"number": season_num, "episodes": episodes})
        item["seasons"] = seasons
        shows.append(item)

    return {"movies": movies, "shows": shows}


def build_ratings_payloads(plan: dict[str, Any], dataset: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    """Construit les payloads POST /sync/ratings (note 0-100)."""
    sections = dataset.get("sections") or {}
    ratings = sections.get("ratings") or {}
    movies, shows = [], []
    for row in ratings.get("movies") or []:
        if not isinstance(row, dict):
            continue
        movie = row.get("movie") if isinstance(row.get("movie"), dict) else row
        ids = movie.get("ids") if isinstance(movie.get("ids"), dict) else {}
        if not _has_mapping(ids):
            continue
        try:
            score = int(round(float(row.get("rating") or 0) * 10))
        except (TypeError, ValueError):
            continue
        item = _ids_block(ids)
        item["rating"] = max(0, min(score, 100))
        movies.append(item)
    for row in ratings.get("shows") or []:
        if not isinstance(row, dict):
            continue
        show = row.get("show") if isinstance(row.get("show"), dict) else row
        ids = show.get("ids") if isinstance(show.get("ids"), dict) else {}
        if not _has_mapping(ids):
            continue
        try:
            score = int(round(float(row.get("rating") or 0) * 10))
        except (TypeError, ValueError):
            continue
        item = _ids_block(ids)
        item["rating"] = max(0, min(score, 100))
        shows.append(item)
    return {"movies": movies, "shows": shows}


def build_watchlist_payloads(dataset: dict[str, Any]) -> dict[str, list[dict[str, Any]]]:
    """Payloads POST /watchlist/items/add."""
    sections = dataset.get("sections") or {}
    watchlist = sections.get("watchlist") or {}
    movies, shows = [], []
    for item in watchlist.get("movies") or []:
        ids = item.get("ids") if isinstance(item.get("ids"), dict) else {}
        if _has_mapping(ids):
            movies.append(_ids_block(ids))
    for item in watchlist.get("shows") or []:
        ids = item.get("ids") if isinstance(item.get("ids"), dict) else {}
        if _has_mapping(ids):
            shows.append(_ids_block(ids))
    return {"movies": movies, "shows": shows}


def build_lists_plans(dataset: dict[str, Any]) -> list[dict[str, Any]]:
    """Plan de création des listes : nom + items (sans correspondance exclus)."""
    sections = dataset.get("sections") or {}
    user_lists = sections.get("user_lists") or []
    out = []
    for item in user_lists:
        if not isinstance(item, dict):
            continue
        name = str(item.get("name") or "Liste")
        movies = []
        for m in item.get("movies") or []:
            ids = m.get("ids") if isinstance(m.get("ids"), dict) else {}
            if _has_mapping(ids):
                movies.append(_ids_block(ids))
        shows = []
        for s in item.get("shows") or []:
            ids = s.get("ids") if isinstance(s.get("ids"), dict) else {}
            if _has_mapping(ids):
                shows.append(_ids_block(ids))
        out.append({"name": name, "movies": movies, "shows": shows})
    return out


# ── Rapport Excel ────────────────────────────────────────────────────────────


def _style_sheet(ws) -> None:
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=f"Tab_{ws.title.replace(' ', '_')}", ref=ref)
        table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
        ws.add_table(table)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="00524B", end_color="00524B", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")


def generate_migration_report(
    plan: dict[str, Any],
    results: dict[str, Any] | None = None,
) -> bytes:
    """Génère le rapport Excel de migration (Résumé, Historique, Notes,
    Watchlist, Listes, Sans correspondance, Échecs)."""
    results = results or {}
    buffer = io.BytesIO()

    # Résumé
    resume = [
        ["Films vus migrés", plan["films_vus"]],
        ["Épisodes vus migrés", plan["episodes_vus"]],
        ["Séries concernées", plan["series_vues"]],
        ["Rewatches (MDBList ne garde que la dernière date)", plan["rewatches"]],
        ["Notes films", plan["notes"]["movies"]],
        ["Notes séries", plan["notes"]["shows"]],
        ["Watchlist films", plan["watchlist"]["movies"]],
        ["Watchlist séries", plan["watchlist"]["shows"]],
        ["Listes à créer", len(plan["listes"])],
        ["Contenus sans correspondance", len(plan["sans_correspondance"])],
    ]
    if results:
        resume.append(["Historique écrit (films)", results.get("watched_movies_ok", 0)])
        resume.append(["Historique écrit (épisodes)", results.get("watched_episodes_ok", 0)])
        resume.append(["Échecs", results.get("errors", 0)])
    df_resume = pd.DataFrame(resume, columns=["Indicateur", "Valeur"])

    # Historique (films)
    hist_films = pd.DataFrame([
        {"Type": "Film", "Titre": m.get("title"), "Année": m.get("year"),
         "Vues": m.get("plays", 1), "Date dernière vue": m.get("watched_at") or "—"}
        for m in plan["movies_seen"].values()
    ])

    # Historique (épisodes par série)
    hist_eps = []
    for entry in plan["episodes_by_show"].values():
        for season in sorted(entry["seasons"]):
            for number, date_str in sorted(entry["seasons"][season].items()):
                hist_eps.append({
                    "Série": entry["title"], "Année": entry.get("year"),
                    "Saison": season, "Épisode": number,
                    "Date vue": date_str or "—",
                })
    df_eps = pd.DataFrame(hist_eps)

    # Sans correspondance
    df_nomatch = pd.DataFrame([
        {"Type": n.get("type"), "Titre": n.get("title"), "Année": n.get("year"), "Raison": n.get("reason")}
        for n in plan["sans_correspondance"]
    ])

    # Échecs
    errors = results.get("errors_list") or []
    df_errors = pd.DataFrame([
        {"Section": e.get("section"), "Détail": e.get("detail")} for e in errors
    ])

    # Watchlist / Listes (info)
    df_watchlist = pd.DataFrame([
        {"Type": "Film" if k == "movies" else "Série", "Quantité": v}
        for k, v in plan["watchlist"].items()
    ])
    listes_detail = plan.get("lists_detail") or [
        {"name": l.get("name"), "movies": [], "shows": []} for l in plan.get("listes") or []
    ]
    df_lists = pd.DataFrame([
        {"Liste": l["name"], "Films": len(l["movies"]), "Séries": len(l["shows"])}
        for l in listes_detail
    ])

    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        df_resume.to_excel(writer, sheet_name="Résumé", index=False)
        hist_films.to_excel(writer, sheet_name="Historique films", index=False)
        df_eps.to_excel(writer, sheet_name="Historique épisodes", index=False)
        df_nomatch.to_excel(writer, sheet_name="Sans correspondance", index=False)
        df_watchlist.to_excel(writer, sheet_name="Watchlist", index=False)
        df_lists.to_excel(writer, sheet_name="Listes", index=False)
        df_errors.to_excel(writer, sheet_name="Échecs", index=False)
    buffer.seek(0)

    wb = load_workbook(buffer)
    for sheet in wb:
        _style_sheet(sheet)
    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final.getvalue()
