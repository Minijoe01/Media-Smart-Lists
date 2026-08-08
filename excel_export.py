"""Exports Excel multi-onglets de Media Smart Lists.

Reprend le principe de l'ancienne application Trakt Smart Lists : chaque
analyse correspond à un onglet du classeur (Résumé, Historique, Watchlist,
Listes, Statistiques, Badges, …), avec entêtes verts et tableaux formatés.
"""

from __future__ import annotations

import io
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def _ajuster(ws) -> None:
    for cell in ws[1]:
        width = min(max(len(str(cell.value or "")) + 2, 8), 60)
        ws.column_dimensions[get_column_letter(cell.column)].width = width


def _forme(ws, coul: str = "00524B") -> None:
    """En-tête coloré + tableau strié + ligne figée, comme Trakt Smart Lists."""
    ws.freeze_panes = "A2"
    if ws.max_row > 1:
        ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        table = Table(displayName=f"Tab_{ws.title.replace(' ', '_')}", ref=ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        ws.add_table(table)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color=coul, end_color=coul, fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    _ajuster(ws)


def build_excel(
    summary_rows: list[tuple[str, Any]],
    history_df: pd.DataFrame,
    watchlist_df: pd.DataFrame,
    lists_df: pd.DataFrame,
    stats_df: pd.DataFrame,
    achievements_df: pd.DataFrame,
) -> bytes:
    """Construit un classeur Excel multi-onglets et le retourne en octets."""
    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        pd.DataFrame(summary_rows, columns=["Statistique", "Valeur"]).to_excel(
            writer, sheet_name="Résumé", index=False
        )
        history_df.to_excel(writer, sheet_name="Historique", index=False)
        watchlist_df.to_excel(writer, sheet_name="Watchlist", index=False)
        lists_df.to_excel(writer, sheet_name="Listes", index=False)
        stats_df.to_excel(writer, sheet_name="Statistiques", index=False)
        achievements_df.to_excel(writer, sheet_name="Badges", index=False)
    buffer.seek(0)

    workbook = load_workbook(buffer)
    for sheet in workbook:
        _forme(sheet)
    # Mise en évidence de la colonne « % » dans Statistiques si présente.
    ws_stats = workbook["Statistiques"]
    pct_col = None
    for cell in ws_stats[1]:
        if cell.value and "%" in str(cell.value):
            pct_col = cell.column
    if pct_col:
        letter = get_column_letter(pct_col)
        if ws_stats.max_row > 1:
            ws_stats.conditional_formatting.add(
                f"{letter}2:{letter}{ws_stats.max_row}",
                ColorScaleRule(
                    start_type="min", start_color="63BE7B",
                    mid_type="percentile", mid_value=50, mid_color="FFEB84",
                    end_type="max", end_color="F8696B",
                ),
            )
    final = io.BytesIO()
    workbook.save(final)
    final.seek(0)
    return final.getvalue()
